import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import time

def format_excel_file(filename, payout_df, processed_df):
    """Applies detailed formatting to the generated Excel file."""
    
    # Use ExcelWriter to write both dataframes to the same file
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        if not payout_df.empty:
            payout_df.to_excel(writer, sheet_name='Payout', index=False)
        if not processed_df.empty:
            processed_df.to_excel(writer, sheet_name='Processed', index=False)
    
    # Re-open the file with openpyxl to apply detailed formatting
    workbook = openpyxl.load_workbook(filename)

    # --- Define styles ---
    header_font = Font(bold=True)
    total_row_font = Font(bold=True)
    header_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    
    # Define border styles
    thin_side = Side(style='thin')
    
    accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        if sheet.max_row == 0: # Skip empty sheets
            continue
            
        # --- Header Formatting ---
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # --- Auto-adjust column widths ---
        for col in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # --- Total Row Formatting ---
        if sheet.max_row > 1:
            total_row_index = sheet.max_row
            for cell in sheet[total_row_index]:
                cell.font = total_row_font
                if cell.value: # Only fill cells that have content (e.g., "Total" and the sum)
                    cell.fill = header_fill
        
        # --- Apply Borders ---
        max_col = sheet.max_column
        max_row = sheet.max_row

        # 1. Header block outer border
        if max_row >= 1:
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col_idx)
                border = Border(top=thin_side, bottom=thin_side)
                if col_idx == 1:
                    border.left = thin_side
                if col_idx == max_col:
                    border.right = thin_side
                cell.border = border
        
        # 2. Main data area outer border
        if max_row > 2: # Check if there is at least one data row
            data_min_row = 2
            data_max_row = max_row - 1
            for row in sheet.iter_rows(min_row=data_min_row, max_row=data_max_row, min_col=1, max_col=max_col):
                for cell in row:
                    border = Border()
                    if cell.row == data_min_row:
                        border.top = thin_side
                    if cell.row == data_max_row:
                        border.bottom = thin_side
                    if cell.column == 1:
                        border.left = thin_side
                    if cell.column == max_col:
                        border.right = thin_side
                    cell.border = border
        
        # 3. Total row cells with values
        if max_row > 1:
            for cell in sheet[max_row]:
                if cell.value:
                    cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


        # --- Dynamic Number Formatting ---
        headers = [cell.value for cell in sheet[1]]
        
        cols_to_format = []
        if sheet_name == 'Payout':
            cols_to_format = ["Payment Amount", "Coupa Customer Dividend"]
        elif sheet_name == 'Processed':
            cols_to_format = ["Payment Amount"]

        for col_name in cols_to_format:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for cell in sheet[col_letter]:
                    if cell.row > 1:
                        cell.number_format = accounting_format

    workbook.save(filename)


def main():
    """Main function to run the data processing script."""
    start_time = time.time()

    # --- Configuration ---
    base_path = r"G:\Shared drives\Partner   Usage\Coupa Pay\Paymode X\Individual Reports\Paymode Python file"
    input_filename = os.path.join(base_path, "Coupa Paymode-X Dividends Report.xlsx")
    output_folder = os.path.join(base_path, "Px Customers Breakdown")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    try:
        xls = pd.ExcelFile(input_filename)
        payout_df_full = pd.read_excel(xls, sheet_name="Payout")
        processed_df_full = pd.read_excel(xls, sheet_name="Processed")
    except FileNotFoundError:
        print(f"Error: The file '{input_filename}' was not found.")
        return
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return

    # Clean column names by stripping whitespace
    payout_df_full.columns = payout_df_full.columns.str.strip()
    processed_df_full.columns = processed_df_full.columns.str.strip()

    # --- NEW: Rename columns as requested ---
    rename_dict = {
        'Disburser Paymode-X Account': 'Disburser Paymode Account',
        'Collector Paymode-X Account': 'Collector Paymode'
    }
    payout_df_full.rename(columns=rename_dict, inplace=True)
    processed_df_full.rename(columns=rename_dict, inplace=True)

    # --- UPDATED: Define desired columns with new names ---
    payout_cols_desired = [
        "Disburser Company Name", "Disburser Paymode Account", 
        "Collector Paymode", "Collector Network Fee Billing Method", 
        "Channel Dividend Currency", "DPA", "Payment Credit Settlement Date", 
        "Date Fees Collected", "Payment Amount", "Coupa Customer Dividend","Payment Number"
    ]
    processed_cols_desired = [
        "Disburser Company Name", "Disburser Paymode Account", 
        "Collector Paymode", "Collector Network Fee Billing Method", 
        "Channel Dividend", "Currency", "DPA", "Payment Credit Settlement Date", 
        "Date Fees Collected", "Payment Amount", "Fee Details","Payment Number"
    ]
    
    # Validate and Filter Columns to prevent KeyErrors
    actual_payout_cols = payout_df_full.columns.tolist()
    payout_cols_to_keep = [col for col in payout_cols_desired if col in actual_payout_cols]
    missing_payout_cols = [col for col in payout_cols_desired if col not in actual_payout_cols]
    if missing_payout_cols:
        print(f"Warning: The following columns were not found in the 'Payout' sheet and will be skipped: {missing_payout_cols}")
    payout_df_full = payout_df_full[payout_cols_to_keep]

    actual_processed_cols = processed_df_full.columns.tolist()
    processed_cols_to_keep = [col for col in processed_cols_desired if col in actual_processed_cols]
    missing_processed_cols = [col for col in processed_cols_desired if col not in actual_processed_cols]
    if missing_processed_cols:
        print(f"Warning: The following columns were not found in the 'Processed' sheet and will be skipped: {missing_processed_cols}")
    processed_df_full = processed_df_full[processed_cols_to_keep]

    all_customers = pd.concat([
        payout_df_full["Disburser Company Name"],
        processed_df_full["Disburser Company Name"]
    ]).unique()

    for customer in all_customers:
        if pd.isna(customer):
            continue

        print(f"Processing customer: {customer}")

        customer_payout_df = payout_df_full[payout_df_full["Disburser Company Name"] == customer].copy()
        customer_processed_df = processed_df_full[processed_df_full["Disburser Company Name"] == customer].copy()

        if not customer_payout_df.empty:
            payout_totals = {"Disburser Company Name": "Total"}
            if "Payment Amount" in customer_payout_df.columns:
                payout_totals["Payment Amount"] = customer_payout_df["Payment Amount"].sum()
            if "Coupa Customer Dividend" in customer_payout_df.columns:
                payout_totals["Coupa Customer Dividend"] = customer_payout_df["Coupa Customer Dividend"].sum()
            payout_total_row = pd.DataFrame([payout_totals])
            customer_payout_df = pd.concat([customer_payout_df, payout_total_row], ignore_index=True)

        if not customer_processed_df.empty:
            processed_totals = {"Disburser Company Name": "Total"}
            if "Payment Amount" in customer_processed_df.columns:
                processed_totals["Payment Amount"] = customer_processed_df["Payment Amount"].sum()
            processed_total_row = pd.DataFrame([processed_totals])
            customer_processed_df = pd.concat([customer_processed_df, processed_total_row], ignore_index=True)
            
        safe_customer_name = "".join(c for c in customer if c.isalnum() or c in (' ', '.', '_')).rstrip()
        output_filename = os.path.join(output_folder, f"{safe_customer_name} Monthly Dividend Report.xlsx")
        
        format_excel_file(output_filename, customer_payout_df, customer_processed_df)

    end_time = time.time()
    total_time = end_time - start_time
    minutes = int(total_time // 60)
    seconds = int(total_time % 60)

    print("\n--- Script Finished ---")
    print(f"Total execution time: {minutes} minutes and {seconds} seconds.")


if __name__ == "__main__":
    main()